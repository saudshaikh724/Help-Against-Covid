from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep
from openpyxl import load_workbook
import pymongo

client = pymongo.MongoClient("mongodb+srv://saud:mlab3431@hospitals.5loco.mongodb.net/helpagainstcovid?retryWrites=true&w=majority")

mydatabase = client.helpagainstcovid

wb=load_workbook("dataq.xlsx")
ws = wb.active
name_Column = ws['A']
address_Column = ws['C']
phone_Column = ws['B']
latlong_Column = ws['D']

PATH = "chromedriver.exe"
op = webdriver.ChromeOptions()
op.add_argument('headless')
driver = webdriver.Chrome(options=op)
driver.implicitly_wait(8)
driver.get("https://mumgis.mcgm.gov.in/portal/apps/opsdashboard/index.html#/bf879b53a6964365bc694adb0cceb4bf")

sleep(3)
btn=driver.find_elements_by_xpath('.//*[@class="btn btn-white toggle-btn"]')
btn[0].click()
sleep(3)
ele = driver.find_elements_by_xpath('.//*[@class="side-nav dropdown-menu"]/a')
st = ""
word1 ='Total Beds'
word2 ='Suspect'
word3 ='All Quarantine'
word4 = 'Centers'
iso = 'All Isolation'
sleep(3)
hplst = []
for i in range(1,150):
    print(i)
    btn[1].click()
    sleep(3)
    if True:
        ele[i].click()
        sleep(3)

        for j in driver.find_elements(By.TAG_NAME, "body"):
            st+=j.text
        st = st.replace('\n', ' ')
        index1 = st.center(len(st) + 2, ' ').find(word1.center(len(word1) + 2, ' '))
        index2 = st.center(len(st) + 2, ' ').find(word2.center(len(word2) + 2, ' '))
        index3 = st.center(len(st) + 2, ' ').find(word3.center(len(word3) + 2, ' '))
        index4 = st.center(len(st) + 2, ' ').find(word4.center(len(word4) + 2, ' '))

        st1=st[index1:index2]
        totalbeds=st1[11:13]
        availbeds=st1[18:]
        sub = st[index3:index4]

        no = []
        length = len(sub) - 1
        abc = 0
        while abc <= length:
            if sub[abc].isdigit() and not sub[abc + 1].isdigit():
                no.append(sub[abc])

            elif sub[abc].isdigit() and sub[abc + 1].isdigit():
                no.append(sub[abc] + sub[abc + 1])
                abc += 1
            abc += 1
        print(no)
        temp = "<table border=1><tr><td>Total Beds</td><td>Quarantine Beds</td><td>Isolation Beds</td><td>ICU Beds</td><td>PICU Beds</td><td>NICU Beds</td><td>Pregnant Beds</td><td>Other Beds</td></tr><tr><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td></tr><tr><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td></tr></table>".format(totalbeds,no[0], no[2], no[4], no[6],no[8], no[10], no[12], availbeds,no[1],no[3], no[5], no[7], no[9], no[11], no[13])
        test = { "name": name_Column[i].value,"hpinfo": address_Column[i].value+"<br><br> Phone No: "+str(phone_Column[i].value),"latlong":latlong_Column[i].value,"bedinfo":temp}
        rec = mydatabase.abc.insert_one(test)
        st=""
